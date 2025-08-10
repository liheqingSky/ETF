import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook

import os
import fnmatch

from Logger import log


class Trade():
    def __init__(self, filename):
        if filename == "":
            self._account_dict = {}
            self._stock_info = []
        else:
            account_dict, stocks_info = self.parse(filename)
            self._account_dict = account_dict
            self._stock_info = stocks_info

    def get_stocks_info(self):
        return (self._account_dict, self._stock_info)

    def parse(self, filename):
        if filename.endswith("xlsx"):
            return self.parse_new(filename)
        elif filename.endswith("xls"):
            return self.parse_old(filename)
        else:
            log.error("not support this format! filename: {}".format(filename))
            return ()

    def parse_new(self, filename):
        workbook = load_workbook(filename)
        worksheet = workbook.active

        total_rows = worksheet.max_row
        total_cols = worksheet.max_column
        log.debug("总行数：{}".format(total_rows))
        log.debug("总列数：{}".format(total_cols))

        row = 1
        account_dict = {}
        for col in range(1, 6):
            key = worksheet.cell(row=row, column=col).value
            value = worksheet.cell(row=row + 1, column=col).value
            account_dict[key] = value
        account_dict = self.data_clean_dict(account_dict)

        # bugfix: 修复银河格式转换华宝，资产计算不对的问题
        if float(account_dict['资产']) == 0:
            account_dict['资产'] = str(float(account_dict['可用']) + float(account_dict['参考市值']))
        log.debug(account_dict)

        key_row = 4
        stocks_info = []
        for row in range(key_row + 1, total_rows + 1):
            current_stock_info = {}
            for col in range(1, total_cols):
                key = worksheet.cell(row=key_row, column=col).value
                value = worksheet.cell(row=row, column=col).value
                current_stock_info[key] = value
            log.debug(current_stock_info)
            stocks_info.append(current_stock_info)

        account_dict = self.data_clean_dict(account_dict)
        stocks_info = self.data_clean_list(stocks_info)
        return account_dict, stocks_info

    def data_clean_dict(self, dicts):
        if isinstance(dicts, dict) == False:
            log.error("var:{} is not dict".format(dicts))
            return

        new_dict = {}
        for key, value in dicts.items():
            new_key = str(key)
            new_value = str(value)
            # log.debug("old key:{} value:{} type: {} {}".format(key, value, type(key), type(value)))
            if key.startswith("="):
                new_key = key[1:]
            if isinstance(value, str) and value.startswith("="):
                new_value = value[1:]

            new_key = new_key.replace('"', '')
            new_value = new_value.replace('"', '')
            # log.debug("new key:{} value:{} type: {} {}".format(new_key, new_value, type(new_key), type(new_value)))
            new_dict[new_key] = new_value


        return new_dict

    def data_clean_list(self, lists):
        new_list = []
        if isinstance(lists, list) == False:
            log.error("var:{} is not list".format(lists))
            return new_list

        for _list in lists:
            new_list.append(self.data_clean_dict(_list))
        return new_list

    def get_excel_format_is_huabao(self, filename):
        element_count = 0
        if filename.endswith("xlsx"):
            workbook = load_workbook(filename)
            worksheet = workbook.active
            first_row = worksheet[1]
            element_count = len(first_row)
        elif filename.endswith("xls"):
            workbook = xlrd.open_workbook(filename)
            worksheet = workbook.sheet_by_index(0)
            first_row = worksheet.row_values(0)
            element_count = len(first_row)

        log.debug("第一行元素个数: {}".format(element_count))
        if element_count == 14:
            return True
        elif element_count == 15:
            return False
        return True

    def save_merge(self, filename):
        print("{} {}".format(self._account_dict, self._stock_info))
        self.write_excel(filename, self._account_dict, self._stock_info)

    def parse_old(self, filename):
        workbook = xlrd.open_workbook(filename)
        worksheet = workbook.sheet_by_index(0)

        account_dict = {}
        for col in range(0, 6):
            account_dict[worksheet.cell_value(0, col)] = worksheet.cell_value(1, col)
        log.debug(account_dict)

        # 获取总行数和列数
        total_rows = worksheet.nrows
        total_cols = worksheet.ncols
        log.debug("总行数：{}".format(total_rows))
        log.debug("总列数：{}".format(total_cols))

        # 解析股票数据汇总为列表
        stocks_info = []
        for row in range(4, total_rows):
            current_stock_info = {}
            for col in range(0, total_cols):
                current_stock_info[worksheet.cell_value(3, col)] = worksheet.cell_value(row, col)
            log.debug(current_stock_info)
            stocks_info.append(current_stock_info)
        log.debug("stocks: {}".format(stocks_info))
        account_dict = self.data_clean_dict(account_dict)
        stocks_info = self.data_clean_list(stocks_info)
        return (account_dict, stocks_info)

    def merge_stock_info(self, stock_1, stock_2):
        merge_stock = {}
        no_need_merge = ['证券代码', '证券名称', '当前价', '摊簿成本价', '参考盈亏比例(%)', '股东代码', '交易所名称']
        for key, value in stock_1.items():
            if key not in no_need_merge:
                merge_stock[key] = round(float(stock_1[key]) + float(stock_2[key]), 3)
                # log.debug("key: {} value: {}".format(key, merge_stock[key]))
            else:
                merge_stock[key] = value

        # log.debug("{} {} {}".format(type(merge_stock['最新市值']), type(merge_stock['摊簿浮动盈亏']), type(merge_stock['证券数量'])))
        merge_stock['摊簿成本价'] = round(
            (float(merge_stock['最新市值']) - float(merge_stock['摊簿浮动盈亏'])) / float(merge_stock['证券数量']), 3)
        merge_stock['参考盈亏比例(%)'] = round(
            float(merge_stock['摊簿浮动盈亏']) / (float(merge_stock['最新市值']) - float(merge_stock['摊簿浮动盈亏'])) * 100, 3)
        return merge_stock

    def merge_stocks_info(self, stock_1, stock_2):
        if stock_1 == None and stock_2 == None:
            log.error("stock1 and stock2 all none")
            return {}

        if stock_2 == None:
            log.debug("stock2 none")
            return stock_1

        if stock_1 == None:
            log.debug("stock_1 none")
            return stock_2

        stocks_info = []
        stocks_info.extend(stock_1)
        stocks_info.extend(stock_2)

        # for stock_info in stocks_info:
        #     log.debug(stock_info)

        # 记录重复的股票列表索引值
        duplicate_info = {}
        list_index = 0
        for stock_info in stocks_info:
            log.debug(stock_info)
            code = stock_info['证券代码']
            if code not in duplicate_info.keys():
                duplicate_info[code] = [list_index]
            else:
                duplicate_info[code].append(list_index)
            list_index += 1

        # log.debug(duplicate_info)

        # 重复的股票汇总
        new_stocks_info = []
        has_merge = {}
        for stock_info in stocks_info:
            code = stock_info['证券代码']
            if len(duplicate_info[code]) == 1:
                new_stocks_info.append(stock_info)
            elif len(duplicate_info[code]) > 1:
                if code in has_merge.keys():
                    continue

                stock_info_1 = stocks_info[duplicate_info[code][0]]
                stock_info_2 = stocks_info[duplicate_info[code][1]]

                # log.debug("before merge {} {}".format(stock_info_1, stock_info_2))
                merge_stock_info = self.merge_stock_info(stock_info_1, stock_info_2)
                # log.debug("after merge {} ".format(merge_stock_info))
                new_stocks_info.append(merge_stock_info)
                has_merge[code] = 1

        for stock_info in new_stocks_info:
            log.debug(stock_info)
        return new_stocks_info

    def get_yinhe_stocks_info(self, filename):
        if filename.endswith("xlsx"):
            workbook = load_workbook(filename)
            worksheet = workbook.active

            total_rows = worksheet.max_row
            total_cols = worksheet.max_column
            log.debug("总行数：{}".format(total_rows))
            log.debug("总列数：{}".format(total_cols))

            stocks_info = []
            for row in range(2, total_rows+1):
                current_stock_info = {}
                for col in range(1, total_cols+1):
                    key = worksheet.cell(row=1, column=col).value
                    value = worksheet.cell(row=row, column=col).value
                    current_stock_info[key] = value
                log.debug("yinhe_stock_info: {}".format(current_stock_info))
                stocks_info.append(current_stock_info)
            log.debug("stocks: {}".format(stocks_info))
            return stocks_info
        elif filename.endswith("xls"):
            workbook = xlrd.open_workbook(filename)
            worksheet = workbook.sheet_by_index(0)

            # 获取总行数和列数
            total_rows = worksheet.nrows
            total_cols = worksheet.ncols
            stocks_info = []
            for row in range(1, total_rows):
                current_stock_info = {}
                for col in range(0, total_cols):
                    current_stock_info[worksheet.cell_value(0, col)] = worksheet.cell_value(row, col)
                log.debug("yinhe_stock_info: {}".format(current_stock_info))
                stocks_info.append(current_stock_info)
            # log.debug("stocks: {}".format(stocks_info))
            return stocks_info
        return {}

    def data_convert(self, yinhe_data):
        # 字典按照指定的顺序进行排列
        order = ['证券代码', '证券名称', '证券数量', '可卖数量', '摊簿成本价', '当前价', '最新市值', '摊簿浮动盈亏',
                 '实现盈亏', '参考盈亏比例(%)', '冻结数量', '非流通数量', '股东代码', '交易所名称']
        new_stocks_info = []
        for stock_info in yinhe_data:
            log.debug(stock_info)
            new_stock_info = {}
            for key, value in stock_info.items():
                if key == '当前持仓':
                    new_stock_info['证券数量'] = value
                elif key == '可用余额':
                    new_stock_info['可卖数量'] = value
                elif key == '参考成本价':
                    new_stock_info['摊簿成本价'] = value
                elif key == '参考市价':
                    new_stock_info['当前价'] = value
                elif key == '参考市值':
                    new_stock_info['最新市值'] = value
                elif key == '买入冻结':
                    new_stock_info['冻结数量'] = value
                elif key == '参考盈亏':
                    new_stock_info['摊簿浮动盈亏'] = value
                elif key == '盈亏比例(%)':
                    new_stock_info['参考盈亏比例(%)'] = value
                elif key == '证券代码':
                    new_stock_info['证券代码'] = value
                elif key == '证券名称':
                    new_stock_info['证券名称'] = value
                elif key == '股东代码':
                    new_stock_info['股东代码'] = value
                elif key == '交易市场':
                    if value == '深Ａ':
                        new_stock_info['交易所名称'] = "上海A股"
                    elif value == '沪Ａ':
                        new_stock_info['交易所名称'] = "深圳A股"

                # 默认补齐
                new_stock_info['非流通数量'] = 0
                new_stock_info['实现盈亏'] = 0
                new_stock_info = dict(zip(order, new_stock_info.values()))
            new_stocks_info.append(new_stock_info)

        account_dict = {'币种': '人民币', '余额': 0, '可用': 0, '参考市值': 0, '资产': 0, '盈亏': 0}
        for new_stock_info in new_stocks_info:
            account_dict['参考市值'] += round(float(new_stock_info['最新市值']), 3)
            account_dict['盈亏'] += round(float(new_stock_info['摊簿浮动盈亏']), 3)

        return (account_dict, new_stocks_info)

    def yinhe_convert_to_huabao(self, filename, new_filename):
        old_stocks_info = self.get_yinhe_stocks_info(filename)
        account_dict, new_stocks_info = self.data_convert(old_stocks_info)
        self.write_excel(new_filename, account_dict, new_stocks_info)

    def write_excel(self, filename, account_dict, stocks_info):
        workbook = Workbook()
        worksheet = workbook.active

        # 写入股票汇总信息
        col = 1
        for key, value in account_dict.items():
            # 获取单元格的行和列
            worksheet.cell(row=1, column=col, value=key)
            worksheet.cell(row=2, column=col, value=value)
            col += 1

        # 写入单只股票持仓信息
        row = 4
        for stock_info in stocks_info:
            col = 1
            for key, value in stock_info.items():
                # 股票列信息只写一次
                if row == 4:
                    worksheet.cell(row=row, column=col, value=key)
                worksheet.cell(row=row + 1, column=col, value=value)
                col += 1
            row += 1
        workbook.save(filename)

    def data_format(self, data):
        return round(float(data), 3)

    def is_all_digits(self, s):
        try:
            float(s)
            return True
        except ValueError:
            return False

    def merge(self, filename):
        account_dict, stocks_info = self.parse(filename)
        # log.debug("account: {} old: {}".format(account_dict, self._account_dict))
        # log.debug("stock_info: {} old: {}".format(stocks_info, self._stock_info))
        # merge account
        new_account = {}
        for key, value in account_dict.items():
            # log.debug("key:{} value:{} type:{}, {}".format(key, value, type(key), type(value)))
            if self.is_all_digits(value) == True:
                # log.debug("{} key:{} value:{} account：{}".format(filename, key, value, self._account_dict))
                old_value = self._account_dict.get(key, 0)
                new_account[key] = self.data_format(float(value) + float(old_value))
            else:
                new_account[key] = value

        # merge stocks info
        new_stock_info = self.merge_stocks_info(stocks_info, self._stock_info)
        self._stock_info = new_stock_info
        self._account_dict = new_account
        log.debug("account: {}".format(self._account_dict))
        log.debug("stock_info: {}".format(self._stock_info))

    def get_stock_info(self, code):
        query_info = {}
        for stock_info in self._stock_info:
            log.debug(stock_info)
            if int(stock_info['证券数量']) == int(code):
                query_info = stock_info
                log.debug("find spec stock info! code: {}, info: {}".format(code, stock_info))
                break
        return query_info

    def get_stock_codes(self):
        codes = []
        for stock_info in self._stock_info:
            code = int(stock_info['证券数量'])
            # log.debug("证券代码: {}".format(code))
            codes.append(code)
        return codes

    def find_xls_files(self, directory):
        xls_files = []
        for dirpath, dirnames, filenames in os.walk(directory):
            for filename in fnmatch.filter(filenames, '*.xlsx'):
                xls_files.append(os.path.join(dirpath, filename))
        return xls_files

    def merge_path(self, path="C:\\Users\\hi\\Documents\\"):
        xls_files = self.find_xls_files(path)
        log.debug(xls_files)
        for xls_file in xls_files:
            log.debug(xls_file)
            if self.get_excel_format_is_huabao(xls_file) == False:
                log.debug("yinhe->huabao")
                directory = os.path.dirname(xls_file)
                filename = os.path.basename(xls_file)
                new_filename = "new" + filename
                new_xls_file = os.path.join(directory, new_filename)
                self.yinhe_convert_to_huabao(xls_file, new_xls_file)
                self.merge(new_xls_file)
            elif xls_file.startswith("new") == False: # 去除银河转华宝的历史记录
                log.debug("huabao")
                self.merge(xls_file)
        self.save_merge("merge.xlsx")

if __name__ == '__main__':
    instance = Trade("")
    instance.merge_path()

